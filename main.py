from selenium import webdriver
from openpyxl import Workbook, load_workbook
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
import time

try:
    wb = Workbook()
    wb.save('vehicles.xlsx')
    ws = wb.create_sheet('Vehicles')
finally:
    wb = load_workbook('vehicles.xlsx')
    ws = wb.create_sheet('Vehicles')


path = "G:/Project/outlook_account_automation/chromedriver.exe"
args = ["hide_console"]
driver = webdriver.Chrome("G:/Project/Gta_base_scraper/chromedriver.exe", service_args=args)

GTA_base_website = driver.get("https://www.gtabase.com/grand-theft-auto-v/vehicles/")

time.sleep(10)

list_of_links = []
condition = True
while condition:
    products = driver.find_elements_by_xpath("//a[@class='product-item-link']")
    for product in products:
        list_of_links.append(product.get_attribute('href'))
    driver.execute_script("window.scrollTo(0, 500);")
    WebDriverWait(driver, 20).until(
        ec.element_to_be_clickable((By.XPATH, "//a[@class='page action next']"))).click()
    try:
        WebDriverWait(driver, 20).until(
            ec.element_to_be_clickable((By.XPATH, "//a[@title='Next']//parent::li[@class='item pages-item-next']"))).click()
        time.sleep(2)
    except:
        condition = False

#print list_of_links

for link in list_of_links:
    vc_s = ''
    vf_s = ''
    af_s = ''
    driver.get(link)
    name = driver.find_element_by_xpath("//h1[@class='contentheading']").text
    # print name
    vc_list = driver.find_elements_by_xpath("//span[contains(text(),'Vehicle Class')]//following-sibling::span[@class='field-value']//span//a")
    for vc in vc_list:
        vc_s += vc.text + ', '
    # print vc_s
    vf_list = driver.find_elements_by_xpath("//span[contains(text(),'Vehicle Features')]//following-sibling::span[@class='field-value']//span//a")
    for vf in vf_list:
        vf_s += vf.text + ', '
    # print vf_s
    af_list = driver.find_elements_by_xpath("//span[contains(text(),'Available from')]//following-sibling::span[@class='field-value']//span//a")
    for af in af_list:
        af_s += af.text + ', '
    # print af_s
    try:
        price = driver.find_element_by_xpath("//span[contains(text(),'GTA Online Price')]//following-sibling::span[@class='field-value']").text
    except:
        price = 0
    # print price
    try:
        real = driver.find_element_by_xpath("//span[contains(text(),'Based on (Real Life)')]//following-sibling::span[@class='field-value']").text
    except:
        real = 'none'
    time.sleep(2)
    ws.append([name, vc_s, price, vf_s, af_s, real])
    time.sleep(2)
    # print price
wb.save('vehicles.xlsx')
driver.close()

